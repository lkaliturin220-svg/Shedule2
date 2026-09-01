"""
Парсер файлов расписания формата studentam_*.xlsx

Поддерживает два формата:

Формат A (будни) — 2 подгруппы, блок 6 колонок:
  +0,+1: дисциплина пг1 (merged)
  +2:    дисциплина пг2
  +3:    кабинет пг2
  +4,+5: кабинет пг1 (merged)
  Строка ниже: преподаватели в тех же смещениях

Формат B (суббота/3 подгруппы) — 3 подгруппы, блок 6 колонок:
  +1: дисциплина пг1,  +2: кабинет пг1
  +3: дисциплина пг2,  +4: кабинет пг2
  +5: (пусто или пг3)
  Признак: строка после заголовков содержит [1, 2, 3] в блочных позициях
"""
import re
from datetime import datetime
from typing import Optional

import openpyxl

BLOCK_STARTS = [1, 7, 13, 19]   # начало каждого блока группы (0-индекс)

GROUP_RE = re.compile(r"^[0-9А-ЯA-Zа-яa-z].{0,18}[-–]\d")

PAIR_TIMES = {
    1: "08:30–10:00",
    2: "10:20–11:50",
    3: "12:10–13:40",
    4: "14:00–15:30",
    5: "15:40–17:10",
    6: "17:15–18:45",
    7: "18:20–19:50",
    8: "20:00–21:30",
}


def _cell(row: tuple, idx: int) -> str:
    if idx >= len(row):
        return ""
    v = row[idx]
    if v is None:
        return ""
    return str(v).strip()


def _is_group_row(row: tuple) -> list[Optional[str]]:
    """Если строка содержит имена групп в блочных позициях — вернуть список."""
    found = []
    has_any = False
    for col in BLOCK_STARTS:
        v = _cell(row, col)
        if v and GROUP_RE.match(v):
            found.append(v)
            has_any = True
        else:
            found.append(None)
    return found if has_any else []


def _is_subgroup_row(row: tuple) -> bool:
    """Строка с номерами подгрупп: col[0]=None, а в блочных позициях стоят 1,2,3."""
    # Первая колонка должна быть пустой (не номер пары)
    if row and row[0] is not None:
        return False
    count = 0
    for col in BLOCK_STARTS:
        for offset in (0, 2, 4):
            v = _cell(row, col + offset)
            if v in ("1", "2", "3"):
                count += 1
    # Минимум 3 совпадения чтобы не путать с обычными строками
    return count >= 3


def _normalize_room(raw: str) -> str:
    if not raw or raw in ("None", "null"):
        return ""
    raw = raw.strip()
    if raw.lower() in ("maх", "max", "мах", "макс", "макс."):
        return "Макс."
    if raw.lower() in ("сам.изуч", "сам. изуч.", "самост", "сам.изуч."):
        return "Сам.изуч."
    return raw


def _detect_format(rows: list) -> str:
    """
    Определить формат файла:
    - 'B' если есть строка с номерами подгрупп [1,2,3] (суббота/3 подгруппы)
    - 'A' иначе (стандартный будничный, 2 подгруппы)
    """
    for row in rows[:15]:
        if _is_subgroup_row(row):
            return "B"
    return "A"


def _parse_block_A(row, teacher_row, col) -> tuple:
    """
    Формат A (будни):
    +0,+1: дисциплина пг1; +2: дисциплина пг2; +3: кабинет пг2; +4,+5: кабинет пг1
    """
    subj1 = _cell(row, col + 0) or _cell(row, col + 1)
    subj2 = _cell(row, col + 2)
    room1 = _normalize_room(_cell(row, col + 4) or _cell(row, col + 5))
    room2 = _normalize_room(_cell(row, col + 3))

    teach1 = teach2 = ""
    if teacher_row is not None:
        teach1 = _cell(teacher_row, col + 0) or _cell(teacher_row, col + 1)
        teach2 = _cell(teacher_row, col + 2) or _cell(teacher_row, col + 3)

    return subj1, room1, teach1, subj2, room2, teach2


def _parse_block_B(row, teacher_row, col) -> tuple:
    """
    Формат B (суббота/3 подгруппы):
    +0: дисциплина пг1;  +1: кабинет пг1
    +2: дисциплина пг2;  +3: кабинет пг2
    +4: кабинет (целая группа, если нет подгрупп)
    """
    subj1 = _cell(row, col + 0)
    room1 = _normalize_room(_cell(row, col + 1))
    subj2 = _cell(row, col + 2)
    room2 = _normalize_room(_cell(row, col + 3))
    # Кабинет целой группы (если нет подгрупп)
    room_single = _normalize_room(_cell(row, col + 4) or _cell(row, col + 5))

    teach1 = teach2 = ""
    if teacher_row is not None:
        teach1 = _cell(teacher_row, col + 0)
        teach2 = _cell(teacher_row, col + 2)

    # Если только целая группа — кабинет берём из +4
    if subj1 and not subj2 and room_single:
        room1 = room_single

    return subj1, room1, teach1, subj2, room2, teach2


def parse_xlsx(file_path_or_bytes) -> dict:
    """
    Возвращает словарь:
    {
        "date": "10.03.2026",
        "day_of_week": "Вторник",
        "lessons": [
            {
                "date": date,
                "day_of_week": str,
                "pair_number": int,
                "subgroup": int | None,
                "group": str,
                "teacher": str,
                "subject": str,
                "room": str,
            },
            ...
        ]
    }
    """
    wb = openpyxl.load_workbook(file_path_or_bytes, read_only=True, data_only=True)

    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    day_of_week = sheet_name.strip()

    # Дата из первой ячейки
    title = _cell(rows[0], 0) if rows else ""
    date_match = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", title)
    if date_match:
        parsed_date = datetime.strptime(date_match.group(0), "%d.%m.%Y").date()
        date_str = date_match.group(0)
    else:
        from datetime import date as d_today
        parsed_date = d_today.today()
        date_str = parsed_date.strftime("%d.%m.%Y")

    # Определяем формат
    fmt = _detect_format(rows)

    lessons = []
    current_groups: list[Optional[str]] = [None, None, None, None]

    i = 0
    while i < len(rows):
        row = rows[i]

        # Пустая строка
        if all(v is None for v in row):
            i += 1
            continue

        # Строка с номерами подгрупп — пропускаем
        if _is_subgroup_row(row):
            i += 1
            continue

        # Строка с именами групп
        group_row = _is_group_row(row)
        if group_row:
            current_groups = group_row
            i += 1
            continue

        # Строка с номером пары
        pair_num_raw = row[0] if row else None
        if pair_num_raw is not None and str(pair_num_raw).strip().isdigit():
            pair_num = int(str(pair_num_raw).strip())
            if 1 <= pair_num <= 8:
                # Следующая строка — преподаватели?
                teacher_row = None
                if i + 1 < len(rows):
                    next_row = rows[i + 1]
                    nv0 = next_row[0] if next_row else None
                    if (nv0 is None or not str(nv0).strip().isdigit()) and not _is_group_row(next_row):
                        teacher_row = next_row

                for block_idx, col in enumerate(BLOCK_STARTS):
                    group_name = current_groups[block_idx] if block_idx < len(current_groups) else None
                    if not group_name:
                        continue

                    if fmt == "B":
                        subj1, room1, teach1, subj2, room2, teach2 = _parse_block_B(row, teacher_row, col)
                    else:
                        subj1, room1, teach1, subj2, room2, teach2 = _parse_block_A(row, teacher_row, col)

                    if subj1 and subj2:
                        lessons.append(_make(parsed_date, day_of_week, pair_num, 1,
                                             group_name, teach1, subj1, room1))
                        lessons.append(_make(parsed_date, day_of_week, pair_num, 2,
                                             group_name, teach2, subj2, room2))
                    elif subj1:
                        lessons.append(_make(parsed_date, day_of_week, pair_num, None,
                                             group_name, teach1, subj1, room1 or room2))
                    elif subj2:
                        lessons.append(_make(parsed_date, day_of_week, pair_num, 2,
                                             group_name, teach2, subj2, room2))

                i += 2 if teacher_row is not None else 1
                continue

        i += 1

    return {
        "date": date_str,
        "day_of_week": day_of_week,
        "lessons": lessons,
    }


def _make(date, dow, pair, subgroup, group, teacher, subject, room) -> dict:
    return {
        "date":        date,
        "day_of_week": dow,
        "pair_number": pair,
        "subgroup":    subgroup,
        "group":       group,
        "teacher":     teacher.strip() if teacher else "",
        "subject":     subject.strip(),
        "room":        room,
    }
